from openpyxl import Workbook,load_workbook
import openpyxl



def criar_planilha():
    arquivo = Workbook()

    planilha=arquivo.active
    planilha.title='Potencia'
 
    plan_THD= arquivo.create_sheet('THD')   
    plan_Variação_de_Tensão= arquivo.create_sheet('Variação')   
    plan_Interrupção=arquivo.create_sheet('Interrupção')  
    plan_Grandezas=arquivo.create_sheet('Grandezas')  
    plan_FFT=arquivo.create_sheet('FFT')
    plan_Freq=arquivo.create_sheet('Frequência')



    planilha['A1']='Potencia Aparente Fase A (VA)'
    planilha['B1']='Potencia Aparente Fase B (VA)'
    planilha['C1']='Potencia Aparente Fase C (VA)'
    planilha['D1']='Potencia Ativa Fase A (w)'
    planilha['E1']='Potencia Ativa Fase B (w)'
    planilha['F1']='Potencia Ativa Fase C (w)'
    planilha['G1']='Potencia Reativa Fase A (Var)'
    planilha['H1']='Potencia Reativa Fase B (Var)'
    planilha['I1']='Potencia Reativa Fase C (Var)'


    plan_THD['A1']='THD Tensão Fase A (%)'
    plan_THD['B1']='THD Tensão Fase B (%)'
    plan_THD['C1']='THD Tensão Fase C (%)'
    plan_THD['D1']='THD Corrente Fase A (%)'
    plan_THD['E1']='THD Corrente Fase B (%)'
    plan_THD['F1']='THD Corrente Fase C (%)'

    plan_Variação_de_Tensão['A1']='Fase A'
    plan_Variação_de_Tensão['B1']='Fase B'
    plan_Variação_de_Tensão['C1']='Fase C'

    plan_Interrupção['A1']='Fase A'
    plan_Interrupção['B1']='Fase B'
    plan_Interrupção['C1']='Fase C'

    plan_Grandezas['A1']='Tensão Fase A (V)'
    plan_Grandezas['B1']='Tensão Fase B (V)'
    plan_Grandezas['C1']='Tensão Fase C (V)'
    plan_Grandezas['D1']='Corrente Fase A (A)'
    plan_Grandezas['E1']='Corrente Fase B (A)'
    plan_Grandezas['F1']='Corrente Fase C (A)'
    plan_Grandezas['G1']='FP Fase A'
    plan_Grandezas['H1']='FP Fase B'
    plan_Grandezas['I1']='FP Fase C'

    
    plan_FFT['A1']='FREQ FFT Tensão Fase A'
    plan_FFT['B1']='FFT Tensão Fase A'
    plan_FFT['C1']='FREQ FFT Tensão Fase B'
    plan_FFT['D1']='FFT Tensão Fase B'
    plan_FFT['E1']='FREQ FFT Tensão Fase C'
    plan_FFT['F1']='FFT Tensão Fase C'
    plan_FFT['G1']='FREQ FFT Corrente Fase A'
    plan_FFT['H1']='FFT Crorrente Fase A'
    plan_FFT['I1']='FREQ FFT Corrente Fase B'
    plan_FFT['J1']='FFT Crorrente Fase B'
    plan_FFT['K1']='FREQ FFT Corrente Fase B'
    plan_FFT['L1']='FFT Crorrente Fase B'

    plan_Freq['A1']='Fase A'
    plan_Freq['B1']='Fase B'
    plan_Freq['C1']='Fase C'
    
    arquivo.save('Dados_Analisador.xlsx')



def armazenar_excel(dado,nome_arquivo,nome_planilha,coluna):

    arquivo=openpyxl.load_workbook(nome_arquivo)
    planilha=arquivo[nome_planilha]

    for i,valor in enumerate(dado,start=2):
            planilha[f'{coluna}{i}'] = valor

    arquivo.save(nome_arquivo)

def Retornar_excel(nome_arquivo,nome_planilha,coluna):
    arquivo=openpyxl.load_workbook(nome_arquivo)
    planilha=arquivo[nome_planilha]
    i=2
    vetor=[]
    while(True):
        x=planilha[f'{coluna}{i}'].value
        if x is None or x == "":
            break
        vetor.append(x)
        i+=1
    
    return vetor
